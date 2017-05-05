import openpyxl
wb = openpyxl.load_workbook('bse.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

import plotly.plotly as py
from plotly.tools import FigureFactory as FF
from datetime import datetime

import pandas.io.data as web

df = web.DataReader("bse", datetime(2003,6, 23), datetime(2017,1,13))
fig = FF.create_candlestick(df.Open, df.High, df.Low, df.Close, dates=df.index)
py.iplot(fig, filename='finance/aapl-candlestick', validate=False)




i = 0;
cell_date = ""
cell = 0
list_time = []
list1 = []
skip_line = []

for i in range(18714,23627,7):
	skip_line.append(i)
# for i in range(1,23629):
# 	if i in skip_line:
# 		continue
# 	else:
# 		cell_date = str(sheet.cell(row=i, column=1).value)[:11] + str(sheet.cell(row=i, column=2).value)[:8]
# 		list_time.append(cell_date)

for i in range(1,23627):
	if i in skip_line:
		continue
	else:
		cell = sheet.cell(row=i, column=6).value
		list1.append(cell)