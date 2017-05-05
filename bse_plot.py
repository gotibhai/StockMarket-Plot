import openpyxl
import plotly.plotly as py
import plotly.graph_objs as go

from datetime import datetime

wb = openpyxl.load_workbook('bse.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
list1 = []
list2 = []
list_time = []
skip_line = []

cell_date = ""
for i in range(18714,23627,7):
	skip_line.append(i)

for i in range(1,23627):
	if i in skip_line:
		continue
	else:
		cell_date = str(sheet.cell(row=i, column=1).value)[:11] + str(sheet.cell(row=i, column=2).value)[:8]
		list_time.append(cell_date)

list_time = list_time[:-1]

i = 0
cell = 0
for i in range(1,23627):
	if i in skip_line:
		continue
	else:
		cell = sheet.cell(row=i, column=6).value
		list1.append(cell)

period1 = 20
period2 = 50
multiplier1 = 2.0/float(period1 + 1)
multiplier2 = 2.0/float(period2 + 1)

def ema_func(periods,multiplier):
	newlist = []
	j = 0
	ema = 0
	sma = 0
	len_list1 = len(list1);
	while j < periods:
		sma = sma + list1[j]
		j = j + 1
	sma = float(sma/periods)
	i = 0
	while(i < len_list1):
		if(i < periods):
			newlist.append(float(0))
		elif(i == periods):
			newlist.append(float(sma))
		else:
			if list1[i] is not None and newlist[i-1] is not None:
				ema = (list1[i]*multiplier) + (newlist[i-1]*(1 - multiplier))
				newlist.append(ema)
		i = i + 1;
	return newlist

ema_50 = []
ema_50 = ema_func(period2,multiplier2)
ema_20 = []
ema_20 = ema_func(period1,multiplier1)

trace1 = go.Scatter(x=list_time,y=ema_50)
trace2 = go.Scatter(x=list_time,y=ema_20)
trace3 = go.Scatter(x=list_time,y=list1)

data = [trace1 , trace2, trace3]
py.plot(data)